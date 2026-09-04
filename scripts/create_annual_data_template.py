"""Create the blank annual-data Excel exchange template for Li-Yu-Tan.

This stage 2-4A utility only creates a user-fillable exchange workbook. It
does not parse, validate, publish, or activate a formal annual-data version.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


TEMPLATE_VERSION = "2-4A.1"
RESERVOIR_ID = "liyutan"
RESERVOIR_NAME = "鯉魚潭水庫"
SHEET_NAMES = ("版本資訊", "水文Q值", "年度基準出流", "水庫參數")
PERIOD_NAMES = ("上旬", "中旬", "下旬")
CANONICAL_PERIODS = tuple(
    (f"{month:02d}-{period}", month, period)
    for month in range(1, 13)
    for period in PERIOD_NAMES
)
Q_CODES_DESCENDING = tuple(f"q{quantile:02d}_cms" for quantile in range(95, 0, -5))

TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
MACHINE_FILL = PatternFill("solid", fgColor="D9E1F2")
FIXED_FILL = PatternFill("solid", fgColor="E7E6E6")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
WARNING_FILL = PatternFill("solid", fgColor="FCE4D6")
BODY_FONT = Font(name="Microsoft JhengHei", size=10)
SUBHEADER_FONT = Font(name="Microsoft JhengHei", bold=True, color="1F1F1F")
THIN_GRAY = Side(style="thin", color="B7C9D6")
SECTION_BORDER = Border(bottom=Side(style="medium", color="5B9BD5"))
TABLE_BORDER = Border(bottom=THIN_GRAY)
COMMENT_AUTHOR = "User"


def _style_title(ws, end_column: int, title: str) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = ws.cell(1, 1, title)
    cell.fill = TITLE_FILL
    cell.font = Font(name="Microsoft JhengHei", bold=True, color="FFFFFF", size=15)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28
    ws.sheet_view.showGridLines = False


def _style_note(ws, row: int, end_column: int, text: str, *, warning: bool = False) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_column)
    cell = ws.cell(row, 1, text)
    cell.font = BODY_FONT
    cell.fill = WARNING_FILL if warning else SECTION_FILL
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    cell.border = SECTION_BORDER
    ws.row_dimensions[row].height = 34 if warning else 28


def _set_headers(ws, chinese_headers: list[str], machine_headers: list[str]) -> None:
    for column, value in enumerate(chinese_headers, 1):
        cell = ws.cell(4, column, value)
        cell.fill = SECTION_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = TABLE_BORDER
    for column, value in enumerate(machine_headers, 1):
        cell = ws.cell(5, column, value)
        cell.fill = MACHINE_FILL
        cell.font = Font(name="Consolas", bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = TABLE_BORDER
    ws.row_dimensions[4].height = 31
    ws.row_dimensions[5].height = 31


def _add_nonnegative_validation(ws, cell_range: str) -> None:
    validation = DataValidation(
        type="decimal",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=True,
    )
    validation.error = "請輸入大於或等於 0 的數值，或保留空白。"
    validation.errorTitle = "數值格式不正確"
    validation.prompt = "請輸入非負數值；尚無資料時請保留空白。"
    validation.promptTitle = "填寫非負數值"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    ws.add_data_validation(validation)
    validation.add(cell_range)


def _apply_body_style(ws, min_row: int, max_row: int, max_column: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, max_col=max_column):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = TABLE_BORDER


def _build_version_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("版本資訊")
    _style_title(ws, 5, "鯉魚潭水庫年度資料匯入範本－版本資訊")
    _style_note(
        ws,
        2,
        5,
        "本 Excel 為人工填寫與交換格式，不是正式權威資料；完成填寫不代表已發布或啟用年度資料。",
        warning=True,
    )
    _style_note(
        ws,
        3,
        5,
        "本年度實績截止旬以前（含該旬）填本年度實際資料；截止旬以後填前一年度相同旬別資料。這項規則目前適用於「年度基準出流」。",
    )
    headers = ["欄位代碼", "中文名稱", "值", "單位／格式", "填寫說明"]
    for column, value in enumerate(headers, 1):
        cell = ws.cell(4, column, value)
        cell.fill = MACHINE_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = TABLE_BORDER

    fields = (
        ("template_version", "範本版本", TEMPLATE_VERSION, "文字", "固定技術值，請勿修改。", False),
        ("reservoir_id", "水庫識別碼", RESERVOIR_ID, "文字", "固定技術值，請勿修改。", False),
        ("reservoir_name", "水庫名稱", RESERVOIR_NAME, "文字", "固定技術值，請勿修改。", False),
        ("applicable_year", "適用年度", None, "西元年（整數）", "請填四位數西元年，不預填任何年度。", True),
        ("actual_data_cutoff_period", "本年度實績截止旬", None, "固定36旬", "請由下拉選單選取；規則見本表上方說明。", True),
        ("hydrology_source_period", "水文Q值資料來源／統計期間", None, "文字", "請填資料來源、測站、統計期間或可交接索引。", True),
        ("annual_outflow_source", "年度基準出流資料來源", None, "文字", "請填來源文件或可交接索引。", True),
        ("overall_note", "整體資料備註", None, "文字", "請記錄本年度資料範圍、限制或其他重要事項。", True),
    )
    for row, (code, name, value, unit, instruction, editable) in enumerate(fields, 5):
        for column, item in enumerate((code, name, value, unit, instruction), 1):
            cell = ws.cell(row, column, item)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = TABLE_BORDER
            cell.fill = INPUT_FILL if editable and column == 3 else FIXED_FILL
        ws.row_dimensions[row].height = 31

    year_validation = DataValidation(
        type="whole", operator="between", formula1="2000", formula2="2100", allow_blank=True
    )
    year_validation.error = "請輸入 2000 至 2100 之間的四位數西元年，或保留空白。"
    year_validation.prompt = "未決定適用年度時請保留空白。"
    year_validation.showErrorMessage = True
    year_validation.showInputMessage = True
    ws.add_data_validation(year_validation)
    year_validation.add("C8")

    cutoff_validation = DataValidation(
        type="list", formula1="annual_period_keys", allow_blank=True
    )
    cutoff_validation.error = "請從固定36旬下拉清單選取。"
    cutoff_validation.prompt = "截止旬以前含該旬填本年度實際出流，其後填前一年度同旬出流。"
    cutoff_validation.showErrorMessage = True
    cutoff_validation.showInputMessage = True
    ws.add_data_validation(cutoff_validation)
    cutoff_validation.add("C9")
    ws["C8"].comment = Comment("適用年度必須為四位數西元年；此範本刻意不預填年度。", COMMENT_AUTHOR)
    ws["C9"].comment = Comment(
        "本年度實績截止旬以前（含該旬）填本年度實際出流；尚未發生的旬填前一年度相同旬別出流。",
        COMMENT_AUTHOR,
    )
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:E12"
    for column, width in {"A": 31, "B": 30, "C": 34, "D": 18, "E": 54}.items():
        ws.column_dimensions[column].width = width
    ws.print_title_rows = "1:4"


def _build_hydrology_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("水文Q值")
    max_column = 3 + len(Q_CODES_DESCENDING)
    _style_title(ws, max_column, "鯉魚潭水庫年度水文 Q 值（固定36旬）")
    _style_note(
        ws,
        2,
        max_column,
        "Q 值為流量超越機率統計值，單位一律為 cms。黃色欄位由使用者填寫，固定旬別欄位請勿修改。",
    )
    _style_note(
        ws,
        3,
        max_column,
        "每旬應符合 Q5 ≥ Q10 ≥ … ≥ Q95；本表依既有使用習慣由 Q95 排至 Q5。所有業務數值預設保持空白。",
        warning=True,
    )
    chinese_headers = ["固定旬鍵", "月份", "旬別"] + [
        f"Q{quantile}（cms）" for quantile in range(95, 0, -5)
    ]
    _set_headers(ws, chinese_headers, ["period_key", "month", "period", *Q_CODES_DESCENDING])

    for row, (period_key, month, period) in enumerate(CANONICAL_PERIODS, 6):
        ws.cell(row, 1, period_key)
        ws.cell(row, 2, month)
        ws.cell(row, 3, period)
        for column in range(1, max_column + 1):
            ws.cell(row, column).fill = FIXED_FILL if column <= 3 else INPUT_FILL
        for column in range(4, max_column + 1):
            ws.cell(row, column).number_format = "0.###"
    _apply_body_style(ws, 6, 41, max_column)
    _add_nonnegative_validation(ws, f"D6:{get_column_letter(max_column)}41")

    comparisons = [
        f"${get_column_letter(column)}6>${get_column_letter(column + 1)}6"
        for column in range(4, max_column)
    ]
    formula = (
        f"AND(COUNT($D6:${get_column_letter(max_column)}6)=19,"
        f"OR({','.join(comparisons)}))"
    )
    ws.conditional_formatting.add(
        f"D6:{get_column_letter(max_column)}41",
        FormulaRule(formula=[formula], fill=PatternFill("solid", fgColor="F4CCCC")),
    )
    ws["D5"].comment = Comment(
        "q95_cms 表示 Q95 流量，單位 cms。Q 值欄均需為非負數值，且每旬須符合 Q5 ≥ … ≥ Q95。",
        COMMENT_AUTHOR,
    )
    ws.freeze_panes = "D6"
    ws.auto_filter.ref = f"A5:{get_column_letter(max_column)}41"
    for column, width in {"A": 14, "B": 9, "C": 10}.items():
        ws.column_dimensions[column].width = width
    for column in range(4, max_column + 1):
        ws.column_dimensions[get_column_letter(column)].width = 13
    ws.print_title_rows = "1:5"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.zoomScale = 80


def _build_outflow_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("年度基準出流")
    _style_title(ws, 6, "鯉魚潭水庫年度基準出流（固定36旬）")
    _style_note(
        ws,
        2,
        6,
        "本表正式名稱為「年度基準出流」，不是單純的前一年度資料。截止旬以前（含該旬）填本年度實際出流；尚未發生的旬填前一年度相同旬別出流。",
    )
    _style_note(
        ws,
        3,
        6,
        "抗旱、緊急調度及個別推估的臨時調整不直接改入本表，仍由推估流程中的日期覆寫或自訂出流處理。所有業務數值預設保持空白。",
        warning=True,
    )
    chinese_headers = [
        "固定旬鍵",
        "月份",
        "旬別",
        "上灌區需求（cms）",
        "下灌區需求（cms）",
        "公共出水（萬噸／日）",
    ]
    machine_headers = [
        "period_key",
        "month",
        "period",
        "upstream_irrigation_cms",
        "downstream_irrigation_cms",
        "public_water_10k_ton_per_day",
    ]
    _set_headers(ws, chinese_headers, machine_headers)
    for row, (period_key, month, period) in enumerate(CANONICAL_PERIODS, 6):
        ws.cell(row, 1, period_key)
        ws.cell(row, 2, month)
        ws.cell(row, 3, period)
        for column in range(1, 7):
            ws.cell(row, column).fill = FIXED_FILL if column <= 3 else INPUT_FILL
        for column in range(4, 7):
            ws.cell(row, column).number_format = "0.###"
    _apply_body_style(ws, 6, 41, 6)
    _add_nonnegative_validation(ws, "D6:F41")
    ws["D5"].comment = Comment("上灌區需求，單位 cms；請填非負數值。", COMMENT_AUTHOR)
    ws["E5"].comment = Comment("下灌區需求，單位 cms；請填非負數值。", COMMENT_AUTHOR)
    ws["F5"].comment = Comment("公共出水，單位萬噸／日；請填非負數值。", COMMENT_AUTHOR)
    ws.freeze_panes = "D6"
    ws.auto_filter.ref = "A5:F41"
    for column, width in {"A": 14, "B": 9, "C": 10, "D": 27, "E": 29, "F": 34}.items():
        ws.column_dimensions[column].width = width
    ws.print_title_rows = "1:5"


def _build_parameters_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("水庫參數")
    _style_title(ws, 7, "鯉魚潭水庫年度參數")
    _style_note(
        ws,
        2,
        7,
        "正式年度基準變更：未來須建立新的完整年度版本；不得直接覆蓋既有正式年度版本。",
    )
    _style_note(
        ws,
        3,
        7,
        "單次推估的臨時假設：不直接改入年度正式基準，應由個別推估設定處理。所有可填欄位預設保持空白。",
        warning=True,
    )
    chinese_headers = ["參數代碼", "中文名稱", "數值", "單位", "適用起日", "依據／來源", "備註"]
    machine_headers = [
        "parameter_code",
        "parameter_name",
        "value",
        "unit",
        "effective_start_date",
        "source_reference",
        "note",
    ]
    _set_headers(ws, chinese_headers, machine_headers)
    parameters = (
        ("max_capacity_10k_ton", "滿庫容量", "萬噸"),
        ("shilin_ecological_flow_cms", "士林堰生態流量", "cms"),
        ("liyutan_ecological_release_cms", "鯉魚潭最低生態放流量", "cms"),
        ("shilin_diversion_limit_cms", "士林堰引水上限", "cms"),
    )
    for row, (code, name, unit) in enumerate(parameters, 6):
        for column, value in enumerate((code, name, None, unit, None, None, None), 1):
            ws.cell(row, column, value).fill = FIXED_FILL if column in (1, 2, 4) else INPUT_FILL
        ws.cell(row, 3).number_format = "0.###"
        ws.cell(row, 5).number_format = "yyyy-mm-dd"
        ws.row_dimensions[row].height = 31
    _apply_body_style(ws, 6, 9, 7)
    _add_nonnegative_validation(ws, "C6:C9")
    date_validation = DataValidation(
        type="date",
        operator="between",
        formula1="DATE(1900,1,1)",
        formula2="DATE(9999,12,31)",
        allow_blank=True,
    )
    date_validation.error = "請輸入有效日期，或保留空白。"
    date_validation.prompt = "建議使用 yyyy-mm-dd 格式；尚未決定時請保留空白。"
    date_validation.showErrorMessage = True
    date_validation.showInputMessage = True
    ws.add_data_validation(date_validation)
    date_validation.add("E6:E9")
    ws["C5"].comment = Comment(
        "四項參數數值均刻意留白；請依正式業務依據填入非負數值。", COMMENT_AUTHOR
    )
    ws["E5"].comment = Comment(
        "適用起日請填日期；正式基準變更未來須建立新的完整年度版本。", COMMENT_AUTHOR
    )
    ws.freeze_panes = "C6"
    ws.auto_filter.ref = "A5:G9"
    for column, width in {"A": 37, "B": 30, "C": 18, "D": 14, "E": 18, "F": 38, "G": 42}.items():
        ws.column_dimensions[column].width = width
    ws.print_title_rows = "1:5"


def build_workbook() -> Workbook:
    """Return a new blank, styled, user-fillable annual-data workbook."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = "鯉魚潭水庫年度資料匯入範本"
    workbook.properties.subject = "第二階段 2-4A 空白年度資料 Excel 公版"
    workbook.properties.creator = "鯉魚潭水庫庫容推估系統"
    _build_version_sheet(workbook)
    _build_hydrology_sheet(workbook)
    _build_outflow_sheet(workbook)
    _build_parameters_sheet(workbook)
    workbook.defined_names.add(
        DefinedName("annual_period_keys", attr_text="'水文Q值'!$A$6:$A$41")
    )
    return workbook


def write_template(output: str | Path, *, overwrite: bool = False) -> Path:
    """Write the workbook to an explicit path, refusing silent overwrite."""
    output_path = Path(output).expanduser()
    if output_path.suffix.lower() != ".xlsx":
        raise ValueError("輸出檔案必須使用 .xlsx 副檔名。")
    if output_path.exists() and not overwrite:
        raise FileExistsError(f"目標檔案已存在；如需覆蓋請明確加上 --overwrite：{output_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook()
    workbook.save(output_path)
    workbook.close()
    return output_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="建立鯉魚潭水庫 2-4A 空白年度資料 Excel 公版。"
    )
    parser.add_argument(
        "--output",
        required=True,
        type=Path,
        help="明確指定輸出的 .xlsx 檔案；程式不會猜測共享路徑。",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="明確允許覆蓋已存在的輸出檔案。",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        output = write_template(args.output, overwrite=args.overwrite)
    except (FileExistsError, OSError, ValueError) as exc:
        print(f"錯誤：{exc}", file=sys.stderr)
        return 1
    print(f"已建立空白年度資料 Excel 公版：{output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
