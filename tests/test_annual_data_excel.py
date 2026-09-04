import copy
import io
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from annual_data_excel import (
    OLD_VALUE_NOT_RECORDED,
    PARAMETER_METADATA_KEY,
    PREVIEW_NOTICE,
    IssueSeverity,
    candidate_artifacts,
    compare_annual_data,
    parse_annual_data_excel,
)
from scripts.create_annual_data_template import build_workbook
from shared_storage_schema import HYDROLOGY_COLUMNS, OUTFLOW_COLUMNS, Q_COLUMNS, sha256_bytes


def _filled_workbook():
    workbook = build_workbook()
    version = workbook["版本資訊"]
    version["C8"] = 2027
    version["C9"] = "06-中旬"
    version["C10"] = "合成水文來源與統計期間"
    version["C11"] = "合成年度基準出流來源"
    version["C12"] = "純合成測試資料"

    hydrology = workbook["水文Q值"]
    for row in range(6, 42):
        for column, quantile in enumerate(range(95, 0, -5), 4):
            hydrology.cell(row, column, (100 - quantile) + (row - 6) / 100)

    outflow = workbook["年度基準出流"]
    for row in range(6, 42):
        outflow.cell(row, 4, 1 + (row - 6) / 100)
        outflow.cell(row, 5, 2 + (row - 6) / 100)
        outflow.cell(row, 6, 50 + (row - 6) / 10)

    parameters = workbook["水庫參數"]
    for row, value in enumerate((12000, 2.5, 0.4, 32), 6):
        parameters.cell(row, 3, value)
        parameters.cell(row, 5, "2027-01-01")
        parameters.cell(row, 6, f"合成依據 {row}")
        parameters.cell(row, 7, f"合成備註 {row}")
    return workbook


def _workbook_bytes(workbook=None):
    workbook = workbook or _filled_workbook()
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _mutated_bytes(mutator):
    workbook = _filled_workbook()
    mutator(workbook)
    return _workbook_bytes(workbook)


def _with_package_entry(raw, name, content=b"synthetic"):
    stream = io.BytesIO(raw)
    with zipfile.ZipFile(stream, "a") as archive:
        archive.writestr(name, content)
    return stream.getvalue()


def _codes(result):
    return {issue.code for issue in result.issues}


def _current_from_candidate(candidate):
    return {
        "version": {
            "applicable_year": candidate.applicable_year,
            "actual_data_cutoff_period": candidate.actual_data_cutoff_period,
            "hydrology_source_period": candidate.hydrology_source_period,
            "annual_outflow_source": candidate.annual_outflow_source,
            "overall_note": candidate.overall_note,
        },
        "hydrology": copy.deepcopy(candidate.hydrology),
        "outflow_demand": copy.deepcopy(candidate.outflow_demand),
        "reservoir_parameters": copy.deepcopy(candidate.reservoir_parameters),
        PARAMETER_METADATA_KEY: copy.deepcopy(candidate.parameter_metadata),
    }


def test_generated_template_with_synthetic_values_parses_all_four_sheets():
    raw = _workbook_bytes()
    result = parse_annual_data_excel(raw, filename="synthetic.xlsx")

    assert result.ok
    assert result.source_sha256 == sha256_bytes(raw)
    assert result.candidate.source_filename == "synthetic.xlsx"
    assert result.candidate.template_version == "2-4A.1"
    assert result.candidate.reservoir_id == "liyutan"
    assert result.candidate.applicable_year == 2027
    assert result.candidate.actual_data_cutoff_period == "06-中旬"
    assert len(result.candidate.hydrology) == 36
    assert len(result.candidate.hydrology[0]) == len(HYDROLOGY_COLUMNS)
    assert len(result.candidate.outflow_demand) == 36
    assert len(result.candidate.outflow_demand[0]) == len(OUTFLOW_COLUMNS)
    assert set(result.candidate.reservoir_parameters) == {
        "max_capacity_10k_ton",
        "shilin_ecological_flow_cms",
        "liyutan_ecological_release_cms",
        "shilin_diversion_limit_cms",
    }
    assert PREVIEW_NOTICE.startswith("僅供驗證與差異預覽")
    assert set(candidate_artifacts(result.candidate)) == {
        "hydrology_q.csv",
        "outflow_demand.csv",
        "reservoir_parameters.json",
    }


def test_bytesio_position_is_restored_and_explicit_path_is_not_modified(tmp_path):
    raw = _workbook_bytes()
    stream = io.BytesIO(raw)
    stream.seek(17)
    stream_result = parse_annual_data_excel(stream)
    assert stream.tell() == 17
    assert stream_result.ok
    assert stream_result.source_sha256 == sha256_bytes(raw)
    assert stream_result.candidate.source_filename is None

    path = tmp_path / "synthetic.xlsx"
    path.write_bytes(raw)
    before = sha256_bytes(path.read_bytes())
    result = parse_annual_data_excel(path)
    after = sha256_bytes(path.read_bytes())
    assert result.ok
    assert before == after == result.source_sha256
    assert result.candidate.source_filename == "synthetic.xlsx"


def test_source_filename_is_provenance_but_not_business_fingerprint():
    raw = _workbook_bytes()
    unnamed = parse_annual_data_excel(raw)
    named = parse_annual_data_excel(raw, filename="renamed-synthetic.xlsx")

    assert unnamed.ok and named.ok
    assert unnamed.candidate.source_filename is None
    assert named.candidate.source_filename == "renamed-synthetic.xlsx"
    assert unnamed.candidate.fingerprint == named.candidate.fingerprint


@pytest.mark.parametrize("filename", ["input.xlsm", "input.xls", "input.csv"])
def test_only_xlsx_extension_is_accepted(filename):
    result = parse_annual_data_excel(_workbook_bytes(), filename=filename)
    assert not result.ok
    assert "unsupported_extension" in _codes(result)


def test_non_xlsx_bytes_are_rejected_without_traceback():
    result = parse_annual_data_excel(b"not an excel workbook", filename="bad.xlsx")
    assert not result.ok
    assert _codes(result) == {"invalid_xlsx"}


@pytest.mark.parametrize(
    ("entry", "code"),
    [
        ("xl/vbaProject.bin", "macro_not_allowed"),
        ("xl/externalLinks/externalLink1.xml", "external_link_not_allowed"),
    ],
)
def test_macro_content_and_external_links_are_rejected(entry, code):
    result = parse_annual_data_excel(_with_package_entry(_workbook_bytes(), entry))
    assert not result.ok
    assert code in _codes(result)


@pytest.mark.parametrize(
    ("mutator", "code"),
    [
        (lambda wb: wb.remove(wb["年度基準出流"]), "missing_sheet"),
        (lambda wb: wb.create_sheet("額外資料"), "extra_sheet"),
        (lambda wb: wb._sheets.reverse(), "sheet_order_modified"),
    ],
)
def test_fixed_sheet_set_and_order(mutator, code):
    result = parse_annual_data_excel(_mutated_bytes(mutator))
    assert not result.ok
    assert code in _codes(result)


@pytest.mark.parametrize(
    ("cell", "value", "code"),
    [
        ("C5", "unknown", "unsupported_template_version"),
        ("C6", "other", "reservoir_id_mismatch"),
        ("C7", "其他水庫", "reservoir_name_mismatch"),
        ("C8", None, "invalid_applicable_year"),
        ("C8", 1999, "invalid_applicable_year"),
        ("C8", 2101, "invalid_applicable_year"),
        ("C9", "13-上旬", "invalid_cutoff_period"),
        ("C10", "   ", "hydrology_source_missing"),
        ("C11", None, "annual_outflow_source_missing"),
    ],
)
def test_version_information_validation(cell, value, code):
    result = parse_annual_data_excel(
        _mutated_bytes(lambda wb: setattr(wb["版本資訊"][cell], "value", value))
    )
    assert not result.ok
    assert code in _codes(result)


@pytest.mark.parametrize(
    ("sheet", "cell", "value", "code"),
    [
        ("版本資訊", "A8", "changed_code", "field_code_modified"),
        ("水文Q值", "D5", "changed_q", "machine_header_modified"),
        ("水文Q值", "A6", "12-下旬", "period_key_modified"),
        ("水文Q值", "B6", 12, "month_modified"),
        ("水文Q值", "C6", "下旬", "period_modified"),
        ("年度基準出流", "G6", 1, "unknown_data_cell"),
        ("水庫參數", "A10", "extra_parameter", "unknown_data_cell"),
    ],
)
def test_machine_codes_fixed_periods_and_unknown_data_are_rejected(sheet, cell, value, code):
    result = parse_annual_data_excel(
        _mutated_bytes(lambda wb: setattr(wb[sheet][cell], "value", value))
    )
    assert not result.ok
    assert code in _codes(result)


def test_duplicate_and_missing_periods_are_reported_together():
    def mutate(workbook):
        workbook["水文Q值"]["A7"] = "01-上旬"

    result = parse_annual_data_excel(_mutated_bytes(mutate))
    assert not result.ok
    assert {"duplicate_period", "missing_period"} <= _codes(result)


@pytest.mark.parametrize(
    ("value", "code"),
    [
        (None, "required_number_missing"),
        ("文字", "invalid_number"),
        (-1, "negative_number"),
        ("NaN", "invalid_number"),
        ("Infinity", "invalid_number"),
        ("-Infinity", "invalid_number"),
    ],
)
def test_q_values_reject_blank_text_negative_nan_and_infinity(value, code):
    result = parse_annual_data_excel(
        _mutated_bytes(lambda wb: setattr(wb["水文Q值"]["D6"], "value", value))
    )
    assert not result.ok
    assert code in _codes(result)


def test_q_values_are_mapped_from_q95_to_q05_and_equal_neighbors_are_allowed():
    def mutate(workbook):
        sheet = workbook["水文Q值"]
        sheet["V6"] = sheet["U6"].value

    result = parse_annual_data_excel(_mutated_bytes(mutate))
    assert result.ok
    row = result.candidate.hydrology[0]
    assert tuple(field for field in row if field.startswith("q")) == Q_COLUMNS
    assert row["q05_cms"] == row["q10_cms"]


def test_q_order_error_reports_sheet_period_field_and_cell():
    def mutate(workbook):
        workbook["水文Q值"]["V6"] = 1
        workbook["水文Q值"]["U6"] = 2

    result = parse_annual_data_excel(_mutated_bytes(mutate))
    issue = next(issue for issue in result.issues if issue.code == "q_order_invalid")
    assert issue.sheet == "水文Q值"
    assert issue.cell == "U6"
    assert "01-上旬" in issue.message
    assert "q05_cms" in issue.message


@pytest.mark.parametrize("value", [None, "文字", -0.1, "NaN", "Infinity"])
def test_outflow_values_reject_invalid_inputs(value):
    result = parse_annual_data_excel(
        _mutated_bytes(lambda wb: setattr(wb["年度基準出流"]["D6"], "value", value))
    )
    assert not result.ok
    assert _codes(result) & {"required_number_missing", "invalid_number", "negative_number"}


@pytest.mark.parametrize(
    ("cell", "value", "code"),
    [
        ("A6", "other", "missing_parameter"),
        ("A7", "max_capacity_10k_ton", "duplicate_parameter"),
        ("C6", None, "required_number_missing"),
        ("C6", -1, "negative_number"),
        ("D6", "cms", "parameter_unit_invalid"),
        ("E6", "2027/01/01", "invalid_effective_date"),
    ],
)
def test_reservoir_parameter_validation(cell, value, code):
    result = parse_annual_data_excel(
        _mutated_bytes(lambda wb: setattr(wb["水庫參數"][cell], "value", value))
    )
    assert not result.ok
    assert code in _codes(result)


def test_missing_parameter_source_and_note_are_warnings_and_metadata_is_preserved():
    def mutate(workbook):
        workbook["水庫參數"]["F6"] = None
        workbook["水庫參數"]["G7"] = None

    result = parse_annual_data_excel(_mutated_bytes(mutate))
    assert result.ok
    assert all(issue.severity is IssueSeverity.WARNING for issue in result.issues)
    assert {"parameter_source_missing", "parameter_note_missing"} <= _codes(result)
    metadata = result.candidate.parameter_metadata
    assert metadata["max_capacity_10k_ton"]["source_reference"] is None
    assert metadata["shilin_ecological_flow_cms"]["note"] is None
    assert metadata["liyutan_ecological_release_cms"]["effective_start_date"] == "2027-01-01"


@pytest.mark.parametrize(
    ("sheet", "cell"),
    [
        ("版本資訊", "C8"),
        ("水文Q值", "D6"),
        ("年度基準出流", "D6"),
        ("水庫參數", "C6"),
        ("水庫參數", "E6"),
    ],
)
def test_formula_cells_are_rejected(sheet, cell):
    result = parse_annual_data_excel(
        _mutated_bytes(lambda wb: setattr(wb[sheet][cell], "value", "=1+1"))
    )
    assert not result.ok
    assert "formula_not_allowed" in _codes(result)


def test_fingerprint_is_content_stable_and_ignores_workbook_formatting():
    raw = _workbook_bytes()
    left = parse_annual_data_excel(raw).candidate
    workbook = load_workbook(io.BytesIO(raw))
    workbook["水文Q值"].column_dimensions["D"].width = 40
    workbook["水文Q值"]["D6"].fill = copy.copy(workbook["水文Q值"]["E6"].fill)
    right = parse_annual_data_excel(_workbook_bytes(workbook)).candidate
    assert left.fingerprint == right.fingerprint
    assert left.source_sha256 != right.source_sha256


def test_fingerprint_changes_when_semantic_content_changes():
    left = parse_annual_data_excel(_workbook_bytes()).candidate
    right = parse_annual_data_excel(
        _mutated_bytes(lambda wb: setattr(wb["年度基準出流"]["D6"], "value", 9))
    ).candidate
    assert left.fingerprint != right.fingerprint


def test_difference_is_zero_for_identical_content_and_detects_exact_changes():
    candidate = parse_annual_data_excel(_workbook_bytes()).candidate
    current = _current_from_candidate(candidate)
    same = compare_annual_data(candidate, current)
    assert not same.is_first_version
    assert same.total_changes == 0
    assert same.rows("水文Q值") == []
    assert same.rows("水庫參數") == []

    current["hydrology"][0]["q05_cms"] = float(candidate.hydrology[0]["q05_cms"]) - 1
    current["outflow_demand"][1]["upstream_irrigation_cms"] = 0
    current["reservoir_parameters"]["max_capacity_10k_ton"] = 11000
    changed = compare_annual_data(candidate, current)
    assert changed.section_changes["水文Q值"] == 1
    assert changed.section_changes["年度基準出流"] == 1
    assert changed.section_changes["水庫參數"] == 1
    assert changed.rows("水文Q值")[0]["差值"] == pytest.approx(1)


@pytest.mark.parametrize(
    ("metadata_field", "old_value"),
    [
        ("effective_start_date", "2026-12-31"),
        ("source_reference", "另一份合成依據"),
        ("note", "另一段合成備註"),
    ],
)
def test_parameter_metadata_only_change_is_counted(metadata_field, old_value):
    candidate = parse_annual_data_excel(_workbook_bytes()).candidate
    current = _current_from_candidate(candidate)
    parameter_code = "max_capacity_10k_ton"
    current[PARAMETER_METADATA_KEY][parameter_code][metadata_field] = old_value

    difference = compare_annual_data(candidate, current)

    assert difference.total_changes == 1
    assert difference.section_changes["水庫參數"] == 1
    assert difference.rows("水庫參數") == [
        {
            "資料鍵": parameter_code,
            "欄位": metadata_field,
            "舊值": old_value,
            "新值": candidate.parameter_metadata[parameter_code][metadata_field],
            "差值": None,
        }
    ]


def test_numeric_text_parameter_source_is_compared_as_exact_text():
    candidate = parse_annual_data_excel(
        _mutated_bytes(lambda wb: setattr(wb["水庫參數"]["F6"], "value", "115"))
    ).candidate
    current = _current_from_candidate(candidate)
    parameter_code = "max_capacity_10k_ton"
    current[PARAMETER_METADATA_KEY][parameter_code]["source_reference"] = "0115"

    difference = compare_annual_data(candidate, current)

    assert difference.total_changes == 1
    assert difference.rows("水庫參數")[0]["舊值"] == "0115"
    assert difference.rows("水庫參數")[0]["新值"] == "115"


def test_scientific_notation_parameter_note_is_compared_as_exact_text():
    candidate = parse_annual_data_excel(
        _mutated_bytes(lambda wb: setattr(wb["水庫參數"]["G6"], "value", "1000"))
    ).candidate
    current = _current_from_candidate(candidate)
    parameter_code = "max_capacity_10k_ton"
    current[PARAMETER_METADATA_KEY][parameter_code]["note"] = "1e3"

    difference = compare_annual_data(candidate, current)

    assert difference.total_changes == 1
    assert difference.rows("水庫參數")[0]["舊值"] == "1e3"
    assert difference.rows("水庫參數")[0]["新值"] == "1000"


def test_numeric_text_in_basic_metadata_is_compared_as_exact_text():
    candidate = parse_annual_data_excel(
        _mutated_bytes(lambda wb: setattr(wb["版本資訊"]["C11"], "value", "115"))
    ).candidate
    current = _current_from_candidate(candidate)
    current["version"]["annual_outflow_source"] = "0115"

    difference = compare_annual_data(candidate, current)

    assert difference.total_changes == 1
    assert difference.rows("基本資訊") == [
        {
            "資料鍵": "年度基準",
            "欄位": "年度基準出流來源",
            "舊值": "0115",
            "新值": "115",
            "差值": None,
        }
    ]


def test_csv_numeric_strings_equal_candidate_float_values_without_false_differences():
    candidate = parse_annual_data_excel(_workbook_bytes()).candidate
    current = _current_from_candidate(candidate)
    hydrology_field = "q05_cms"
    outflow_field = "upstream_irrigation_cms"
    current["hydrology"][1][hydrology_field] = str(
        candidate.hydrology[1][hydrology_field]
    )
    current["outflow_demand"][1][outflow_field] = str(
        candidate.outflow_demand[1][outflow_field]
    )

    difference = compare_annual_data(candidate, current)

    assert difference.total_changes == 0


def test_old_version_without_parameter_metadata_is_not_reported_as_identical():
    candidate = parse_annual_data_excel(_workbook_bytes()).candidate
    current = _current_from_candidate(candidate)
    current.pop(PARAMETER_METADATA_KEY)

    difference = compare_annual_data(candidate, current)
    metadata_rows = [
        row for row in difference.rows("水庫參數") if row["欄位"] != "value"
    ]

    assert difference.total_changes == 12
    assert difference.section_changes["水庫參數"] == 12
    assert len(metadata_rows) == 12
    assert {row["舊值"] for row in metadata_rows} == {OLD_VALUE_NOT_RECORDED}


def test_no_active_version_builds_first_complete_preview():
    candidate = parse_annual_data_excel(_workbook_bytes()).candidate
    preview = compare_annual_data(candidate, None)
    assert preview.is_first_version
    assert preview.section_totals == {
        "基本資訊": 5,
        "水文Q值": 36 * 19,
        "年度基準出流": 36 * 3,
        "水庫參數": 4 * 4,
    }
    assert preview.section_changes == preview.section_totals
    assert {
        row["欄位"] for row in preview.rows("水庫參數", changed_only=False)
    } == {"value", "effective_start_date", "source_reference", "note"}
