from pathlib import Path

import pytest
from openpyxl import load_workbook

from scripts.create_annual_data_template import (
    CANONICAL_PERIODS,
    Q_CODES_DESCENDING,
    RESERVOIR_ID,
    RESERVOIR_NAME,
    SHEET_NAMES,
    TEMPLATE_VERSION,
    main,
    write_template,
)


def _create_and_load(tmp_path: Path):
    output = tmp_path / "annual-data-template.xlsx"
    write_template(output)
    return output, load_workbook(output, data_only=False)


def test_xlsx_round_trip_and_sheet_order(tmp_path):
    output, workbook = _create_and_load(tmp_path)
    assert output.is_file()
    assert workbook.sheetnames == list(SHEET_NAMES)
    workbook.close()


def test_fixed_technical_fields_and_blank_business_metadata(tmp_path):
    _, workbook = _create_and_load(tmp_path)
    ws = workbook["版本資訊"]
    metadata = {ws.cell(row, 1).value: ws.cell(row, 3).value for row in range(5, 13)}
    assert metadata == {
        "template_version": TEMPLATE_VERSION,
        "reservoir_id": RESERVOIR_ID,
        "reservoir_name": RESERVOIR_NAME,
        "applicable_year": None,
        "actual_data_cutoff_period": None,
        "hydrology_source_period": None,
        "annual_outflow_source": None,
        "overall_note": None,
    }
    workbook.close()


def test_hydrology_has_canonical_36_periods_and_all_q_columns(tmp_path):
    _, workbook = _create_and_load(tmp_path)
    ws = workbook["水文Q值"]
    headers = [ws.cell(5, column).value for column in range(1, 23)]
    rows = [tuple(ws.cell(row, column).value for column in range(1, 4)) for row in range(6, 42)]
    assert headers == ["period_key", "month", "period", *Q_CODES_DESCENDING]
    assert len(Q_CODES_DESCENDING) == 19
    assert rows == list(CANONICAL_PERIODS)
    assert len(rows) == len(set(rows)) == 36
    assert all(
        ws.cell(row, column).value is None
        for row in range(6, 42)
        for column in range(4, 23)
    )
    workbook.close()


def test_outflow_has_canonical_periods_fields_units_and_blank_values(tmp_path):
    _, workbook = _create_and_load(tmp_path)
    ws = workbook["年度基準出流"]
    assert [ws.cell(5, column).value for column in range(1, 7)] == [
        "period_key",
        "month",
        "period",
        "upstream_irrigation_cms",
        "downstream_irrigation_cms",
        "public_water_10k_ton_per_day",
    ]
    assert [ws.cell(4, column).value for column in range(4, 7)] == [
        "上灌區需求（cms）",
        "下灌區需求（cms）",
        "公共出水（萬噸／日）",
    ]
    rows = [tuple(ws.cell(row, column).value for column in range(1, 4)) for row in range(6, 42)]
    assert rows == list(CANONICAL_PERIODS)
    assert all(
        ws.cell(row, column).value is None
        for row in range(6, 42)
        for column in range(4, 7)
    )
    workbook.close()


def test_reservoir_parameters_are_complete_unique_and_blank(tmp_path):
    _, workbook = _create_and_load(tmp_path)
    ws = workbook["水庫參數"]
    expected_codes = [
        "max_capacity_10k_ton",
        "shilin_ecological_flow_cms",
        "liyutan_ecological_release_cms",
        "shilin_diversion_limit_cms",
    ]
    codes = [ws.cell(row, 1).value for row in range(6, 10)]
    assert codes == expected_codes
    assert len(codes) == len(set(codes))
    for row in range(6, 10):
        assert [ws.cell(row, column).value for column in (3, 5, 6, 7)] == [None] * 4
    workbook.close()


def test_primary_data_validations_and_usability_features_exist(tmp_path):
    _, workbook = _create_and_load(tmp_path)
    version = workbook["版本資訊"]
    hydrology = workbook["水文Q值"]
    outflow = workbook["年度基準出流"]
    parameters = workbook["水庫參數"]
    assert len(version.data_validations.dataValidation) == 2
    assert any(
        validation.type == "whole" and "C8" in str(validation.sqref)
        for validation in version.data_validations.dataValidation
    )
    assert any(
        validation.type == "list"
        and validation.formula1 == "annual_period_keys"
        and "C9" in str(validation.sqref)
        for validation in version.data_validations.dataValidation
    )
    assert "annual_period_keys" in workbook.defined_names
    assert any(
        validation.type == "decimal" and "D6:V41" in str(validation.sqref)
        for validation in hydrology.data_validations.dataValidation
    )
    assert any(
        validation.type == "decimal" and "D6:F41" in str(validation.sqref)
        for validation in outflow.data_validations.dataValidation
    )
    assert any(
        validation.type == "decimal" and "C6:C9" in str(validation.sqref)
        for validation in parameters.data_validations.dataValidation
    )
    assert hydrology.freeze_panes == "D6" and hydrology.auto_filter.ref == "A5:V41"
    assert outflow.freeze_panes == "D6" and outflow.auto_filter.ref == "A5:F41"
    assert parameters.freeze_panes == "C6" and parameters.auto_filter.ref == "A5:G9"
    workbook.close()


def test_missing_output_argument_writes_nothing(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    with pytest.raises(SystemExit) as exc_info:
        main([])
    assert exc_info.value.code == 2
    assert list(tmp_path.iterdir()) == []


def test_existing_output_is_not_silently_overwritten(tmp_path):
    output = tmp_path / "annual-data-template.xlsx"
    output.write_bytes(b"existing-user-content")
    with pytest.raises(FileExistsError, match="--overwrite"):
        write_template(output)
    assert output.read_bytes() == b"existing-user-content"


def test_explicit_overwrite_replaces_existing_file_with_valid_xlsx(tmp_path):
    output = tmp_path / "annual-data-template.xlsx"
    output.write_bytes(b"obsolete")
    write_template(output, overwrite=True)
    workbook = load_workbook(output)
    assert workbook.sheetnames == list(SHEET_NAMES)
    workbook.close()
