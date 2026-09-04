"""Parse, validate, and preview 2-4A annual-data Excel workbooks.

This module is intentionally independent of Streamlit and performs no shared-
storage discovery or writes.  A caller must explicitly provide workbook bytes,
a seekable binary file object, or a local path.
"""

from __future__ import annotations

import datetime as dt
import io
import math
import os
import zipfile
from collections import Counter
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Any, BinaryIO, Iterable, Mapping

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

from shared_storage_schema import (
    HYDROLOGY_COLUMNS,
    OUTFLOW_COLUMNS,
    PERIODS,
    Q_COLUMNS,
    RESERVOIR_PARAMETERS_SCHEMA,
    SCHEMA_VERSION,
    deserialize_csv,
    deterministic_fingerprint,
    serialize_csv,
    serialize_json,
    sha256_bytes,
    validate_hydrology_rows,
    validate_outflow_rows,
    validate_reservoir_parameters,
)
from scripts.create_annual_data_template import (
    CANONICAL_PERIODS,
    Q_CODES_DESCENDING,
    RESERVOIR_ID,
    RESERVOIR_NAME,
    SHEET_NAMES,
    TEMPLATE_VERSION,
)


PREVIEW_NOTICE = "僅供驗證與差異預覽，尚未建立或啟用正式系統基準版本。"
PARAMETER_METADATA_KEY = "parameter_metadata"
OLD_VALUE_NOT_RECORDED = "舊版未記錄"

VERSION_HEADERS = ("欄位代碼", "中文名稱", "值", "單位／格式", "填寫說明")
VERSION_FIELDS = (
    "template_version",
    "reservoir_id",
    "reservoir_name",
    "applicable_year",
    "actual_data_cutoff_period",
    "hydrology_source_period",
    "annual_outflow_source",
    "overall_note",
)
HYDROLOGY_CHINESE_HEADERS = (
    "固定旬鍵",
    "月份",
    "旬別",
    *(f"Q{quantile}（cms）" for quantile in range(95, 0, -5)),
)
HYDROLOGY_MACHINE_HEADERS = ("period_key", "month", "period", *Q_CODES_DESCENDING)
OUTFLOW_CHINESE_HEADERS = (
    "固定旬鍵",
    "月份",
    "旬別",
    "上灌區需求（cms）",
    "下灌區需求（cms）",
    "公共出水（萬噸／日）",
)
OUTFLOW_MACHINE_HEADERS = OUTFLOW_COLUMNS
PARAMETER_CHINESE_HEADERS = (
    "參數代碼",
    "中文名稱",
    "數值",
    "單位",
    "適用起日",
    "依據／來源",
    "備註",
)
PARAMETER_MACHINE_HEADERS = (
    "parameter_code",
    "parameter_name",
    "value",
    "unit",
    "effective_start_date",
    "source_reference",
    "note",
)
PARAMETER_DEFINITIONS = (
    ("max_capacity_10k_ton", "滿庫容量", "萬噸"),
    ("shilin_ecological_flow_cms", "士林堰生態流量", "cms"),
    ("liyutan_ecological_release_cms", "鯉魚潭最低生態放流量", "cms"),
    ("shilin_diversion_limit_cms", "士林堰引水上限", "cms"),
)


class IssueSeverity(str, Enum):
    ERROR = "error"
    WARNING = "warning"


@dataclass(frozen=True)
class AnnualDataIssue:
    severity: IssueSeverity
    code: str
    message: str
    sheet: str | None = None
    cell: str | None = None

    @property
    def location(self) -> str:
        if self.sheet and self.cell:
            return f"{self.sheet}!{self.cell}"
        return self.sheet or self.cell or "整份活頁簿"


@dataclass(frozen=True)
class AnnualDataCandidate:
    template_version: str
    reservoir_id: str
    reservoir_name: str
    applicable_year: int
    actual_data_cutoff_period: str
    hydrology_source_period: str
    annual_outflow_source: str
    overall_note: str | None
    hydrology: tuple[dict[str, Any], ...]
    outflow_demand: tuple[dict[str, Any], ...]
    reservoir_parameters: dict[str, float]
    parameter_metadata: dict[str, dict[str, Any]]
    source_sha256: str
    fingerprint: str
    warnings: tuple[AnnualDataIssue, ...]


@dataclass(frozen=True)
class AnnualDataParseResult:
    source_sha256: str | None
    candidate: AnnualDataCandidate | None
    issues: tuple[AnnualDataIssue, ...]

    @property
    def ok(self) -> bool:
        return self.candidate is not None and not self.errors

    @property
    def errors(self) -> tuple[AnnualDataIssue, ...]:
        return tuple(issue for issue in self.issues if issue.severity is IssueSeverity.ERROR)

    @property
    def warnings(self) -> tuple[AnnualDataIssue, ...]:
        return tuple(issue for issue in self.issues if issue.severity is IssueSeverity.WARNING)


@dataclass(frozen=True)
class DifferenceItem:
    section: str
    data_key: str
    field: str
    old_value: Any
    new_value: Any
    delta: float | None
    changed: bool


@dataclass(frozen=True)
class AnnualDataDifference:
    is_first_version: bool
    items: tuple[DifferenceItem, ...]
    section_totals: dict[str, int]
    section_changes: dict[str, int]

    @property
    def total_changes(self) -> int:
        return sum(self.section_changes.values())

    def rows(self, section: str, *, changed_only: bool = True) -> list[dict[str, Any]]:
        return [
            {
                "資料鍵": item.data_key,
                "欄位": item.field,
                "舊值": item.old_value,
                "新值": item.new_value,
                "差值": item.delta,
            }
            for item in self.items
            if item.section == section and (item.changed or not changed_only)
        ]


def _issue(
    issues: list[AnnualDataIssue],
    severity: IssueSeverity,
    code: str,
    message: str,
    sheet: str | None = None,
    cell: str | None = None,
) -> None:
    issues.append(AnnualDataIssue(severity, code, message, sheet, cell))


def _error(
    issues: list[AnnualDataIssue],
    code: str,
    message: str,
    sheet: str | None = None,
    cell: str | None = None,
) -> None:
    _issue(issues, IssueSeverity.ERROR, code, message, sheet, cell)


def _warning(
    issues: list[AnnualDataIssue],
    code: str,
    message: str,
    sheet: str | None = None,
    cell: str | None = None,
) -> None:
    _issue(issues, IssueSeverity.WARNING, code, message, sheet, cell)


def _read_source(source: bytes | bytearray | BinaryIO | str | os.PathLike[str]) -> tuple[bytes, str | None]:
    if isinstance(source, (str, os.PathLike)):
        path = Path(source)
        return path.read_bytes(), path.name
    if isinstance(source, (bytes, bytearray)):
        return bytes(source), None
    if not hasattr(source, "read"):
        raise TypeError("Excel 輸入必須是 .xlsx bytes、二進位檔案物件或明確 Path。")
    position = None
    if hasattr(source, "tell"):
        try:
            position = source.tell()
        except (OSError, ValueError):
            position = None
    if position is not None and hasattr(source, "seek"):
        try:
            source.seek(0)
        except (OSError, ValueError):
            pass
    data = source.read()
    if position is not None and hasattr(source, "seek"):
        try:
            source.seek(position)
        except (OSError, ValueError):
            pass
    if not isinstance(data, (bytes, bytearray)):
        raise TypeError("Excel 檔案物件必須以二進位模式讀取。")
    return bytes(data), getattr(source, "name", None)


def _has_value(value: Any) -> bool:
    return value is not None and value != ""


def _validate_package(raw: bytes, filename: str | None, issues: list[AnnualDataIssue]) -> bool:
    if filename and Path(filename).suffix.lower() != ".xlsx":
        _error(issues, "unsupported_extension", "只接受不含巨集的 .xlsx 檔案。", cell=filename)
        return False
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            names = set(archive.namelist())
            if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                _error(issues, "invalid_xlsx", "檔案不是可讀取的標準 .xlsx 活頁簿。")
                return False
            content_types = archive.read("[Content_Types].xml").lower()
            if (
                b"macroenabled" in content_types
                or any(name.lower().endswith("vbaproject.bin") for name in names)
            ):
                _error(issues, "macro_not_allowed", "年度基準資料不接受巨集或 .xlsm 內容。")
            if any(name.lower().startswith("xl/externallinks/") for name in names):
                _error(
                    issues,
                    "external_link_not_allowed",
                    "活頁簿含外部連結；請移除外部連結並將業務輸入轉為固定值。",
                )
    except (zipfile.BadZipFile, KeyError, OSError):
        _error(issues, "invalid_xlsx", "檔案不是可讀取的標準 .xlsx 活頁簿。")
        return False
    return not any(issue.severity is IssueSeverity.ERROR for issue in issues)


def _check_extra_values(ws, max_row: int, max_column: int, issues: list[AnnualDataIssue]) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if _has_value(cell.value) and (cell.row > max_row or cell.column > max_column):
                _error(
                    issues,
                    "unknown_data_cell",
                    "固定範本範圍外含有資料，無法安全判斷其用途。",
                    ws.title,
                    cell.coordinate,
                )


def _check_headers(
    ws,
    row: int,
    expected: Iterable[str],
    issues: list[AnnualDataIssue],
    *,
    code: str,
) -> None:
    for column, expected_value in enumerate(expected, 1):
        cell = ws.cell(row, column)
        if cell.value != expected_value:
            _error(
                issues,
                code,
                f"固定表頭應為 {expected_value!r}，目前內容不符。",
                ws.title,
                cell.coordinate,
            )


def _cell_value(cell, issues: list[AnnualDataIssue], *, business_input: bool = True) -> Any:
    if business_input and cell.data_type == "f":
        _error(
            issues,
            "formula_not_allowed",
            "業務輸入不可使用 Excel 公式；請貼上或輸入固定值。",
            cell.parent.title,
            cell.coordinate,
        )
        return None
    return cell.value


def _required_text(cell, issues: list[AnnualDataIssue], code: str, label: str) -> str | None:
    value = _cell_value(cell, issues)
    if not isinstance(value, str) or not value.strip():
        _error(issues, code, f"{label}為必填文字，不可空白。", cell.parent.title, cell.coordinate)
        return None
    return value.strip()


def _optional_text(cell, issues: list[AnnualDataIssue]) -> str | None:
    value = _cell_value(cell, issues)
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if not isinstance(value, str):
        _error(
            issues,
            "invalid_text",
            "此欄位必須是文字。",
            cell.parent.title,
            cell.coordinate,
        )
        return None
    return value.strip()


def _nonnegative_number(cell, issues: list[AnnualDataIssue], field: str) -> float | None:
    value = _cell_value(cell, issues)
    if value is None or value == "":
        _error(
            issues,
            "required_number_missing",
            f"{field}必填，請輸入非負數值。",
            cell.parent.title,
            cell.coordinate,
        )
        return None
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        _error(
            issues,
            "invalid_number",
            f"{field}必須是數字，不接受文字或公式。",
            cell.parent.title,
            cell.coordinate,
        )
        return None
    number = float(value)
    if not math.isfinite(number):
        _error(
            issues,
            "nonfinite_number",
            f"{field}必須是有限數值，不接受 NaN 或 Infinity。",
            cell.parent.title,
            cell.coordinate,
        )
        return None
    if number < 0:
        _error(
            issues,
            "negative_number",
            f"{field}不可為負值。",
            cell.parent.title,
            cell.coordinate,
        )
        return None
    return number


def _validate_fixed_period_rows(ws, issues: list[AnnualDataIssue]) -> None:
    actual_keys: list[str] = []
    expected_keys = {period_key for period_key, _, _ in CANONICAL_PERIODS}
    for row, (period_key, month, period) in enumerate(CANONICAL_PERIODS, 6):
        key_cell, month_cell, period_cell = ws.cell(row, 1), ws.cell(row, 2), ws.cell(row, 3)
        key_value = _cell_value(key_cell, issues)
        month_value = _cell_value(month_cell, issues)
        period_value = _cell_value(period_cell, issues)
        if isinstance(key_value, str) and key_value:
            actual_keys.append(key_value)
        if key_value != period_key:
            _error(
                issues,
                "period_key_modified",
                f"固定旬鍵應為 {period_key}，不得修改。",
                ws.title,
                key_cell.coordinate,
            )
        if isinstance(month_value, float) and month_value.is_integer():
            month_value = int(month_value)
        if month_value != month:
            _error(
                issues,
                "month_modified",
                f"固定月份應為 {month}，不得修改。",
                ws.title,
                month_cell.coordinate,
            )
        if period_value != period:
            _error(
                issues,
                "period_modified",
                f"固定旬別應為 {period}，不得修改。",
                ws.title,
                period_cell.coordinate,
            )
    counts = Counter(actual_keys)
    for key, count in counts.items():
        if count > 1:
            _error(issues, "duplicate_period", f"固定旬鍵 {key} 重複出現 {count} 次。", ws.title, key)
    for key in sorted(expected_keys - set(actual_keys)):
        _error(issues, "missing_period", f"缺少固定旬鍵 {key}。", ws.title, key)
    for key in sorted(set(actual_keys) - expected_keys):
        _error(issues, "unknown_period", f"含未知固定旬鍵 {key}。", ws.title, key)


def _parse_version_sheet(ws, issues: list[AnnualDataIssue]) -> dict[str, Any]:
    _check_extra_values(ws, 12, 5, issues)
    _check_headers(ws, 4, VERSION_HEADERS, issues, code="header_modified")
    actual_codes = [ws.cell(row, 1).value for row in range(5, 13)]
    for offset, expected_code in enumerate(VERSION_FIELDS, 5):
        cell = ws.cell(offset, 1)
        if cell.value != expected_code:
            _error(
                issues,
                "field_code_modified",
                f"固定欄位代碼應為 {expected_code}，不得修改。",
                ws.title,
                cell.coordinate,
            )
    counts = Counter(value for value in actual_codes if isinstance(value, str))
    for code, count in counts.items():
        if count > 1:
            _error(issues, "duplicate_field_code", f"欄位代碼 {code} 重複。", ws.title, code)

    values = {code: _cell_value(ws.cell(row, 3), issues) for row, code in enumerate(VERSION_FIELDS, 5)}
    template_version = values["template_version"]
    if template_version != TEMPLATE_VERSION:
        _error(
            issues,
            "unsupported_template_version",
            f"範本版本必須為 {TEMPLATE_VERSION}；未知版本不可勉強解析。",
            ws.title,
            "C5",
        )
    if values["reservoir_id"] != RESERVOIR_ID:
        _error(issues, "reservoir_id_mismatch", "水庫識別碼必須為 liyutan。", ws.title, "C6")
    if values["reservoir_name"] != RESERVOIR_NAME:
        _error(issues, "reservoir_name_mismatch", "水庫名稱必須為「鯉魚潭水庫」。", ws.title, "C7")

    year = values["applicable_year"]
    if isinstance(year, float) and year.is_integer():
        year = int(year)
    if isinstance(year, bool) or not isinstance(year, int) or not 2000 <= year <= 2100:
        _error(issues, "invalid_applicable_year", "適用年度必須是 2000 至 2100 的四位數西元年。", ws.title, "C8")
        year = None

    cutoff = values["actual_data_cutoff_period"]
    period_keys = {item[0] for item in CANONICAL_PERIODS}
    if cutoff not in period_keys:
        _error(issues, "invalid_cutoff_period", "本年度實績截止旬必須是固定36旬之一。", ws.title, "C9")
        cutoff = None

    hydrology_source = _required_text(ws["C10"], issues, "hydrology_source_missing", "水文Q值資料來源／統計期間")
    outflow_source = _required_text(ws["C11"], issues, "annual_outflow_source_missing", "年度基準出流資料來源")
    overall_note = _optional_text(ws["C12"], issues)
    return {
        "template_version": template_version,
        "reservoir_id": values["reservoir_id"],
        "reservoir_name": values["reservoir_name"],
        "applicable_year": year,
        "actual_data_cutoff_period": cutoff,
        "hydrology_source_period": hydrology_source,
        "annual_outflow_source": outflow_source,
        "overall_note": overall_note,
    }


def _parse_hydrology_sheet(ws, issues: list[AnnualDataIssue]) -> tuple[dict[str, Any], ...]:
    _check_extra_values(ws, 41, len(HYDROLOGY_MACHINE_HEADERS), issues)
    _check_headers(ws, 4, HYDROLOGY_CHINESE_HEADERS, issues, code="header_modified")
    _check_headers(ws, 5, HYDROLOGY_MACHINE_HEADERS, issues, code="machine_header_modified")
    _validate_fixed_period_rows(ws, issues)
    rows: list[dict[str, Any]] = []
    for row_number, (period_key, month, period) in enumerate(CANONICAL_PERIODS, 6):
        q_values: dict[str, float | None] = {}
        row_valid = True
        for column, code in enumerate(Q_CODES_DESCENDING, 4):
            value = _nonnegative_number(ws.cell(row_number, column), issues, f"{period_key} {code}")
            q_values[code] = value
            row_valid = row_valid and value is not None
        if row_valid:
            values = [q_values[code] for code in Q_COLUMNS]
            for index, (left_code, right_code) in enumerate(zip(Q_COLUMNS, Q_COLUMNS[1:])):
                if values[index] < values[index + 1]:
                    right_column = HYDROLOGY_MACHINE_HEADERS.index(right_code) + 1
                    _error(
                        issues,
                        "q_order_invalid",
                        f"{period_key} 必須符合 Q5 ≥ Q10 ≥ … ≥ Q95；{left_code} 不得小於 {right_code}。",
                        ws.title,
                        f"{get_column_letter(right_column)}{row_number}",
                    )
        record: dict[str, Any] = {
            "period_key": period_key,
            "month": month,
            "period": period,
            **{code: q_values[code] for code in Q_COLUMNS},
        }
        rows.append(record)
    return tuple(rows)


def _parse_outflow_sheet(ws, issues: list[AnnualDataIssue]) -> tuple[dict[str, Any], ...]:
    _check_extra_values(ws, 41, len(OUTFLOW_MACHINE_HEADERS), issues)
    _check_headers(ws, 4, OUTFLOW_CHINESE_HEADERS, issues, code="header_modified")
    _check_headers(ws, 5, OUTFLOW_MACHINE_HEADERS, issues, code="machine_header_modified")
    _validate_fixed_period_rows(ws, issues)
    rows: list[dict[str, Any]] = []
    for row_number, (period_key, month, period) in enumerate(CANONICAL_PERIODS, 6):
        record: dict[str, Any] = {"period_key": period_key, "month": month, "period": period}
        for column, code in enumerate(OUTFLOW_COLUMNS[3:], 4):
            record[code] = _nonnegative_number(ws.cell(row_number, column), issues, f"{period_key} {code}")
        rows.append(record)
    return tuple(rows)


def _parse_date(cell, issues: list[AnnualDataIssue]) -> str | None:
    value = _cell_value(cell, issues)
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            parsed = from_excel(value, cell.parent.parent.epoch)
        except (TypeError, ValueError, OverflowError):
            parsed = None
        if isinstance(parsed, dt.datetime):
            return parsed.date().isoformat()
        if isinstance(parsed, dt.date):
            return parsed.isoformat()
    if isinstance(value, str):
        try:
            parsed = dt.date.fromisoformat(value.strip())
        except ValueError:
            parsed = None
        if parsed is not None and value.strip() == parsed.isoformat():
            return parsed.isoformat()
    _error(issues, "invalid_effective_date", "適用起日必須是有效日期，建議格式為 YYYY-MM-DD。", cell.parent.title, cell.coordinate)
    return None


def _parse_parameters_sheet(ws, issues: list[AnnualDataIssue]) -> tuple[dict[str, float], dict[str, dict[str, Any]]]:
    _check_extra_values(ws, 9, len(PARAMETER_MACHINE_HEADERS), issues)
    _check_headers(ws, 4, PARAMETER_CHINESE_HEADERS, issues, code="header_modified")
    _check_headers(ws, 5, PARAMETER_MACHINE_HEADERS, issues, code="machine_header_modified")
    actual_codes = [ws.cell(row, 1).value for row in range(6, 10)]
    counts = Counter(value for value in actual_codes if isinstance(value, str))
    for code, count in counts.items():
        if count > 1:
            _error(issues, "duplicate_parameter", f"參數代碼 {code} 重複。", ws.title, code)
    expected_codes = {item[0] for item in PARAMETER_DEFINITIONS}
    for code in expected_codes - set(actual_codes):
        _error(issues, "missing_parameter", f"缺少參數 {code}。", ws.title, code)
    for code in set(actual_codes) - expected_codes:
        if code is not None:
            _error(issues, "unknown_parameter", f"含未知參數 {code}。", ws.title, str(code))

    parameters: dict[str, float] = {}
    metadata: dict[str, dict[str, Any]] = {}
    for row, (code, name, unit) in enumerate(PARAMETER_DEFINITIONS, 6):
        if ws.cell(row, 1).value != code:
            _error(issues, "parameter_code_modified", f"固定參數代碼應為 {code}。", ws.title, f"A{row}")
        if ws.cell(row, 2).value != name:
            _error(issues, "parameter_name_modified", f"固定參數名稱應為 {name}。", ws.title, f"B{row}")
        if ws.cell(row, 4).value != unit:
            _error(issues, "parameter_unit_invalid", f"{code} 的單位必須為 {unit}。", ws.title, f"D{row}")
        value = _nonnegative_number(ws.cell(row, 3), issues, code)
        if value is not None:
            parameters[code] = value
        effective_date = _parse_date(ws.cell(row, 5), issues)
        source_reference = _optional_text(ws.cell(row, 6), issues)
        note = _optional_text(ws.cell(row, 7), issues)
        if source_reference is None:
            _warning(issues, "parameter_source_missing", f"{code} 尚未填寫依據／來源；發布前請確認。", ws.title, f"F{row}")
        if note is None:
            _warning(issues, "parameter_note_missing", f"{code} 尚未填寫個別備註；本階段仍可預覽。", ws.title, f"G{row}")
        metadata[code] = {
            "parameter_name": name,
            "unit": unit,
            "effective_start_date": effective_date,
            "source_reference": source_reference,
            "note": note,
        }
    return parameters, metadata


def _candidate_payload(
    version: Mapping[str, Any],
    hydrology: Iterable[Mapping[str, Any]],
    outflow: Iterable[Mapping[str, Any]],
    parameters: Mapping[str, float],
    parameter_metadata: Mapping[str, Mapping[str, Any]],
) -> dict[str, Any]:
    return {
        **dict(version),
        "hydrology": [dict(row) for row in hydrology],
        "outflow_demand": [dict(row) for row in outflow],
        "reservoir_parameters": dict(parameters),
        "parameter_metadata": {key: dict(value) for key, value in parameter_metadata.items()},
    }


def candidate_artifacts(candidate: AnnualDataCandidate) -> dict[str, bytes]:
    """Return standard in-memory data bytes without creating a formal version."""
    hydrology_csv = serialize_csv(candidate.hydrology, HYDROLOGY_COLUMNS)
    outflow_csv = serialize_csv(candidate.outflow_demand, OUTFLOW_COLUMNS)
    parameters_json = serialize_json(
        {
            "schema": RESERVOIR_PARAMETERS_SCHEMA,
            "schema_version": SCHEMA_VERSION,
            **candidate.reservoir_parameters,
        }
    )
    validate_hydrology_rows(deserialize_csv(hydrology_csv, HYDROLOGY_COLUMNS))
    validate_outflow_rows(deserialize_csv(outflow_csv, OUTFLOW_COLUMNS))
    validate_reservoir_parameters(
        {
            "schema": RESERVOIR_PARAMETERS_SCHEMA,
            "schema_version": SCHEMA_VERSION,
            **candidate.reservoir_parameters,
        }
    )
    return {
        "hydrology_q.csv": hydrology_csv,
        "outflow_demand.csv": outflow_csv,
        "reservoir_parameters.json": parameters_json,
    }


def parse_annual_data_excel(
    source: bytes | bytearray | BinaryIO | str | os.PathLike[str],
    *,
    filename: str | None = None,
) -> AnnualDataParseResult:
    """Parse one explicitly supplied workbook and return all safe-to-collect issues."""
    issues: list[AnnualDataIssue] = []
    try:
        raw, detected_name = _read_source(source)
    except (OSError, TypeError) as exc:
        _error(issues, "file_read_failed", f"無法讀取 Excel：{exc}")
        return AnnualDataParseResult(None, None, tuple(issues))
    source_sha256 = sha256_bytes(raw)
    filename = filename or detected_name
    if not _validate_package(raw, filename, issues):
        return AnnualDataParseResult(source_sha256, None, tuple(issues))
    try:
        workbook = load_workbook(io.BytesIO(raw), data_only=False, read_only=False, keep_links=True)
    except Exception:
        _error(issues, "invalid_xlsx", "Excel 活頁簿無法安全開啟；請確認檔案未損壞且未加密。")
        return AnnualDataParseResult(source_sha256, None, tuple(issues))
    try:
        actual_sheets = tuple(workbook.sheetnames)
        if actual_sheets != SHEET_NAMES:
            missing = [name for name in SHEET_NAMES if name not in actual_sheets]
            extra = [name for name in actual_sheets if name not in SHEET_NAMES]
            if missing:
                _error(issues, "missing_sheet", f"缺少固定工作表：{'、'.join(missing)}。")
            if extra:
                _error(issues, "extra_sheet", f"含未知工作表：{'、'.join(extra)}。")
            if not missing and not extra:
                _error(issues, "sheet_order_modified", "四張固定工作表的順序不得修改。")
            return AnnualDataParseResult(source_sha256, None, tuple(issues))

        version = _parse_version_sheet(workbook["版本資訊"], issues)
        if any(issue.code == "unsupported_template_version" for issue in issues):
            return AnnualDataParseResult(source_sha256, None, tuple(issues))
        hydrology = _parse_hydrology_sheet(workbook["水文Q值"], issues)
        outflow = _parse_outflow_sheet(workbook["年度基準出流"], issues)
        parameters, parameter_metadata = _parse_parameters_sheet(workbook["水庫參數"], issues)
        if any(issue.severity is IssueSeverity.ERROR for issue in issues):
            return AnnualDataParseResult(source_sha256, None, tuple(issues))

        payload = _candidate_payload(version, hydrology, outflow, parameters, parameter_metadata)
        fingerprint = deterministic_fingerprint(payload)
        warnings = tuple(issue for issue in issues if issue.severity is IssueSeverity.WARNING)
        candidate = AnnualDataCandidate(
            **version,
            hydrology=hydrology,
            outflow_demand=outflow,
            reservoir_parameters=parameters,
            parameter_metadata=parameter_metadata,
            source_sha256=source_sha256,
            fingerprint=fingerprint,
            warnings=warnings,
        )
        try:
            candidate_artifacts(candidate)
        except Exception as exc:
            _error(issues, "schema_validation_failed", f"標準資料結構驗證失敗：{exc}")
            return AnnualDataParseResult(source_sha256, None, tuple(issues))
        return AnnualDataParseResult(source_sha256, candidate, tuple(issues))
    finally:
        workbook.close()


def _baseline_parts(
    current: Any,
) -> tuple[
    dict[str, Any],
    tuple[dict[str, Any], ...],
    tuple[dict[str, Any], ...],
    dict[str, Any],
    dict[str, dict[str, Any]],
]:
    if hasattr(current, "version"):
        parameter_metadata = getattr(current, PARAMETER_METADATA_KEY, {})
        return (
            dict(current.version),
            tuple(dict(row) for row in current.hydrology),
            tuple(dict(row) for row in current.outflow_demand),
            dict(current.reservoir_parameters),
            {
                str(code): dict(metadata)
                for code, metadata in dict(parameter_metadata or {}).items()
                if isinstance(metadata, Mapping)
            },
        )
    if isinstance(current, Mapping):
        parameter_metadata = current.get(PARAMETER_METADATA_KEY, {})
        return (
            dict(current.get("version", {})),
            tuple(dict(row) for row in current.get("hydrology", ())),
            tuple(dict(row) for row in current.get("outflow_demand", ())),
            dict(current.get("reservoir_parameters", {})),
            {
                str(code): dict(metadata)
                for code, metadata in dict(parameter_metadata or {}).items()
                if isinstance(metadata, Mapping)
            },
        )
    raise TypeError("目前年度版本必須是 AnnualDataSnapshot、mapping 或 None。")


def _number_or_original(value: Any) -> Any:
    if isinstance(value, bool) or value is None:
        return value
    try:
        number = float(value)
    except (TypeError, ValueError):
        return value
    return number if math.isfinite(number) else value


def _difference_item(
    section: str,
    key: str,
    field: str,
    old: Any,
    new: Any,
    first: bool,
    *,
    old_recorded: bool = True,
) -> DifferenceItem:
    old_normalized = (
        OLD_VALUE_NOT_RECORDED
        if not first and not old_recorded
        else _number_or_original(old)
    )
    new_normalized = _number_or_original(new)
    changed = True if first or not old_recorded else old_normalized != new_normalized
    delta = None
    if isinstance(old_normalized, (int, float)) and isinstance(new_normalized, (int, float)):
        delta = float(new_normalized) - float(old_normalized)
    return DifferenceItem(section, key, field, old_normalized, new_normalized, delta, changed)


def compare_annual_data(candidate: AnnualDataCandidate, current: Any | None) -> AnnualDataDifference:
    """Compare a candidate with one fully validated active annual snapshot."""
    first = current is None
    if first:
        old_version: dict[str, Any] = {}
        old_hydrology: tuple[dict[str, Any], ...] = ()
        old_outflow: tuple[dict[str, Any], ...] = ()
        old_parameters: dict[str, Any] = {}
        old_parameter_metadata: dict[str, dict[str, Any]] = {}
    else:
        (
            old_version,
            old_hydrology,
            old_outflow,
            old_parameters,
            old_parameter_metadata,
        ) = _baseline_parts(current)
    items: list[DifferenceItem] = []

    metadata_fields = (
        ("applicable_year", "適用年度"),
        ("actual_data_cutoff_period", "實績截止旬"),
        ("hydrology_source_period", "水文資料來源／統計期間"),
        ("annual_outflow_source", "年度基準出流來源"),
        ("overall_note", "整體備註"),
    )
    for field, label in metadata_fields:
        old = old_version.get(field)
        if field == "overall_note" and old is None:
            old = old_version.get("note")
        items.append(_difference_item("基本資訊", "年度基準", label, old, getattr(candidate, field), first))

    old_hydrology_by_key = {row.get("period_key"): row for row in old_hydrology}
    for row in candidate.hydrology:
        old_row = old_hydrology_by_key.get(row["period_key"], {})
        for field in Q_COLUMNS:
            items.append(_difference_item("水文Q值", row["period_key"], field, old_row.get(field), row[field], first))

    old_outflow_by_key = {row.get("period_key"): row for row in old_outflow}
    for row in candidate.outflow_demand:
        old_row = old_outflow_by_key.get(row["period_key"], {})
        for field in OUTFLOW_COLUMNS[3:]:
            items.append(_difference_item("年度基準出流", row["period_key"], field, old_row.get(field), row[field], first))

    for parameter_code, _, _ in PARAMETER_DEFINITIONS:
        items.append(
            _difference_item(
                "水庫參數",
                parameter_code,
                "value",
                old_parameters.get(parameter_code),
                candidate.reservoir_parameters[parameter_code],
                first,
                old_recorded=parameter_code in old_parameters,
            )
        )
        old_metadata = old_parameter_metadata.get(parameter_code, {})
        new_metadata = candidate.parameter_metadata[parameter_code]
        for metadata_field in (
            "effective_start_date",
            "source_reference",
            "note",
        ):
            items.append(
                _difference_item(
                    "水庫參數",
                    parameter_code,
                    metadata_field,
                    old_metadata.get(metadata_field),
                    new_metadata.get(metadata_field),
                    first,
                    old_recorded=metadata_field in old_metadata,
                )
            )

    sections = ("基本資訊", "水文Q值", "年度基準出流", "水庫參數")
    totals = {section: sum(item.section == section for item in items) for section in sections}
    changes = {section: sum(item.section == section and item.changed for item in items) for section in sections}
    return AnnualDataDifference(first, tuple(items), totals, changes)
